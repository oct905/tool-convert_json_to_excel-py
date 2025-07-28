import pandas as pd
import json
import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# Load JSON data from external file
# Change this variable to switch between different JSON files

# json_filename = "report_sync_employee"  

# json_filename = "report_send_course" 
# json_filename = "report_send_event" 
# json_filename = "report_send_event_course" 
json_filename = "report_send_event_participant" 
# json_filename = "report_send_event_participant_evaluation_course" 
# json_filename = "report_send_event_participant_evaluation"

# Check if JSON file exists
json_file_path = json_filename + '.json'
if not os.path.exists(json_file_path):
    print(f"Error: JSON file '{json_file_path}' not found!")
    exit(1)

try:
    with open(json_file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
except json.JSONDecodeError as e:
    print(f"Error: Invalid JSON format in '{json_file_path}': {e}")
    exit(1)
except Exception as e:
    print(f"Error reading file '{json_file_path}': {e}")
    exit(1)

# Check if data is empty
if not data:
    print("Error: JSON file contains no data!")
    exit(1)

# Create DataFrame directly from the cleaned JSON data
df = pd.DataFrame(data)

# Add number column on the left side
df.insert(0, 'No', range(1, len(df) + 1))

# Export to Excel with current date and time in filename
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
output_path = f"./excel/{json_filename}_{current_datetime}.xlsx"

try:
    df.to_excel(output_path, index=False)
except PermissionError:
    print(f"Error: Permission denied. Please close '{output_path}' if it's open in Excel.")
    exit(1)
except Exception as e:
    print(f"Error creating Excel file: {e}")
    exit(1)

# Load the workbook and apply formatting
wb = load_workbook(output_path)
ws = wb.active

# Auto-fit column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    
    # Set minimum width and add some padding
    adjusted_width = min(max_length + 2, 50)  # Max width of 50
    ws.column_dimensions[column_letter].width = adjusted_width

# Create Excel Table with sorting and filtering capabilities
# Define the table range (from A1 to last column and row with data)
last_row = len(df) + 1  # +1 because header is row 1
last_col = len(df.columns)
table_range = f"A1:{ws.cell(row=last_row, column=last_col).coordinate}"

# Create table
table_name = json_filename.replace("-", "_")  # Table names can't contain hyphens
table = Table(displayName=table_name, ref=table_range)

# Add a default table style with alternating row colors
style = TableStyleInfo(
    name="TableStyleMedium9",  # Blue table style
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
table.tableStyleInfo = style

# Add the table to the worksheet
ws.add_table(table)

# Save the formatted workbook
wb.save(output_path)

print(f"Excel file created: {output_path}")
print(f"Total records: {len(df)}")